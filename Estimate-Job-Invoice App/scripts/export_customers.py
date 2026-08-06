"""
export_customers.py — build a private customers.json for the app's "Import customers".

Reads the Clients sheet of client-database.xlsx and writes customers.json next to it.
The app's Import button loads this file so Dad gets repeat-customer autofill on the phone.

PRIVACY: customers.json holds real names, phones, and addresses. Keep it OUT of the
GitHub repo (see .gitignore). Move it to the phone privately (AirDrop / iCloud / Files),
then load it once in the app. It never touches the public web.

Usage:
    python3 scripts/export_customers.py
    python3 scripts/export_customers.py --db /path/to/client-database.xlsx --out customers.json
"""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_DB = Path(__file__).resolve().parent.parent.parent / "client-database.xlsx"

# Clients sheet columns (1-indexed), headers on row 2, data from row 3
COLS = {"first":2,"last":3,"phone":4,"email":5,"street":6,"city":7,"zip":8,"type":9,"status":11}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=str(DEFAULT_DB))
    ap.add_argument("--out", default=str(Path(__file__).resolve().parent.parent / "customers.json"))
    a = ap.parse_args()

    wb = load_workbook(a.db, data_only=True)
    ws = wb["Clients"]
    out = []
    for row in range(3, ws.max_row + 1):
        first = ws.cell(row=row, column=COLS["first"]).value
        last  = ws.cell(row=row, column=COLS["last"]).value
        if not first and not last:
            continue
        cid = "C%03d" % (row - 2)   # matches the sheet's formula-derived Client ID
        def g(k):
            v = ws.cell(row=row, column=COLS[k]).value
            return "" if v is None else str(v).strip()
        out.append({
            "client_id": cid,
            "first": (first or ""), "last": (last or ""),
            "phone": g("phone"), "email": g("email"),
            "street": g("street"), "city": g("city"), "zip": g("zip"),
            "type": g("type") or "Residential",
        })
    Path(a.out).write_text(json.dumps({"clients": out}, indent=2))
    print(f"Wrote {len(out)} customers to {a.out}")
    print("Keep this file private — do not commit it to the repo.")

if __name__ == "__main__":
    main()
