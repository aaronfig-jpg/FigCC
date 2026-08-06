"""
import_jobs.py — write the app's jobs into client-database.xlsx.

Takes the app's JSON export (Saved -> Export JSON) and appends new jobs (and any
new customers) to client-database.xlsx, preserving the sheet's formula-derived IDs.

- New customers (app client_id starting "APP-") are matched to existing rows by phone,
  else appended to Clients; the job then references the real "Cxxx" ID the sheet assigns.
- Existing customers already carry their real "Cxxx" ID from Import customers, so their
  jobs attach straight to that client.
- Already-imported jobs are skipped via a small ledger (scripts/.import_ledger.json).
- A timestamped backup is written to backups/ before any change.

Usage:
    python3 scripts/import_jobs.py figueroas-export.json
    python3 scripts/import_jobs.py figueroas-export.json --db /path/to/client-database.xlsx --dry-run
"""
import argparse, json, shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent.parent           # Figueroa's Carpet Cleaning/
DEFAULT_DB = ROOT / "client-database.xlsx"
BACKUP_DIR = ROOT / "backups"
LEDGER = Path(__file__).resolve().parent / ".import_ledger.json"

CLIENT_HDR_ROW = 2   # data starts row 3
JOB_HDR_ROW = 2

def norm_phone(p): return "".join(ch for ch in str(p or "") if ch.isdigit())

def last_data_row(ws, key_col, start=3):
    r = start - 1
    for row in range(start, ws.max_row + 1):
        if ws.cell(row=row, column=key_col).value not in (None, ""):
            r = row
    return r

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("export", help="app JSON export (figueroas-export.json)")
    ap.add_argument("--db", default=str(DEFAULT_DB))
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    data = json.loads(Path(a.export).read_text())
    jobs = data.get("jobs", [])
    clients = {c["client_id"]: c for c in data.get("clients", [])}
    ledger = set(json.loads(LEDGER.read_text())) if LEDGER.exists() else set()

    wb = load_workbook(a.db)   # keep formulas
    wsc, wsj = wb["Clients"], wb["Jobs"]

    # index existing clients by phone -> real Cxxx
    phone_to_cid = {}
    for row in range(3, wsc.max_row + 1):
        fn = wsc.cell(row=row, column=2).value
        if not fn: continue
        ph = norm_phone(wsc.cell(row=row, column=4).value)
        if ph: phone_to_cid[ph] = "C%03d" % (row - 2)

    c_row = last_data_row(wsc, 2)   # last client by First Name
    j_row = last_data_row(wsj, 2)   # last job by Date
    appid_to_cid = {}
    added_clients = added_jobs = skipped = 0

    def ensure_client(appcid):
        nonlocal c_row, added_clients
        c = clients.get(appcid, {})
        if appcid and not appcid.startswith("APP-"):
            return appcid                      # already a real Cxxx
        if appcid in appid_to_cid:
            return appid_to_cid[appcid]
        ph = norm_phone(c.get("phone"))
        if ph and ph in phone_to_cid:          # matches an existing customer
            appid_to_cid[appcid] = phone_to_cid[ph]; return phone_to_cid[ph]
        c_row += 1                              # append a new client row
        real = "C%03d" % (c_row - 2)
        if not a.dry_run:
            wsc.cell(row=c_row, column=2, value=c.get("first",""))
            wsc.cell(row=c_row, column=3, value=c.get("last",""))
            wsc.cell(row=c_row, column=4, value=c.get("phone",""))
            wsc.cell(row=c_row, column=5, value=c.get("email",""))
            wsc.cell(row=c_row, column=6, value=c.get("street",""))
            wsc.cell(row=c_row, column=7, value=c.get("city",""))
            wsc.cell(row=c_row, column=8, value=c.get("zip",""))
            wsc.cell(row=c_row, column=9, value=c.get("type","Residential"))
            wsc.cell(row=c_row, column=11, value="Active")
            wsc.cell(row=c_row, column=12, value=datetime.today())
        appid_to_cid[appcid] = real
        if ph: phone_to_cid[ph] = real
        added_clients += 1
        return real

    for j in jobs:
        if j.get("job_id") in ledger:
            skipped += 1; continue
        cid = ensure_client((j.get("client") or {}).get("client_id"))
        j_row += 1
        if not a.dry_run:
            wsj.cell(row=j_row, column=2, value=j.get("date"))
            wsj.cell(row=j_row, column=3, value=cid)
            wsj.cell(row=j_row, column=5, value=j.get("service_type"))
            wsj.cell(row=j_row, column=6, value=j.get("areas_rooms_string"))
            wsj.cell(row=j_row, column=7, value=j.get("stain_notes"))
            wsj.cell(row=j_row, column=8, value=j.get("quoted_price"))
            wsj.cell(row=j_row, column=9, value=j.get("discount") or None)
            wsj.cell(row=j_row, column=11, value=j.get("paid"))
            wsj.cell(row=j_row, column=12, value=j.get("payment_method"))
            wsj.cell(row=j_row, column=13, value=j.get("technicians"))
            wsj.cell(row=j_row, column=14, value=j.get("hours_on_site"))
            wsj.cell(row=j_row, column=15, value=j.get("notes"))
        ledger.add(j.get("job_id"))
        added_jobs += 1

    print(f"{'DRY RUN — ' if a.dry_run else ''}jobs added: {added_jobs}, "
          f"new customers: {added_clients}, skipped (already imported): {skipped}")

    if a.dry_run:
        return
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(a.db, BACKUP_DIR / f"client-database-{stamp}.xlsx")
    wb.save(a.db)
    LEDGER.write_text(json.dumps(sorted(ledger)))
    print(f"Saved. Backup: backups/client-database-{stamp}.xlsx")
    print("Tip: open the workbook in Excel/Sheets once so the ID + Final Price formulas recalculate.")

if __name__ == "__main__":
    main()
